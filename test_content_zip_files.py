from pypdf import PdfReader
from download_file import CURRENT_DIR
from openpyxl import load_workbook
import os
import zipfile
import csv
from io import TextIOWrapper

resource_dir = os.path.join(CURRENT_DIR, "resource", "test.zip")


def test_create_archive_1(download_pack_files):
    file_size = os.path.getsize(resource_dir)
    assert file_size == 1175582
    with zipfile.ZipFile(resource_dir, "r") as myzip:
        assert len(myzip.infolist()) == 3


def test_package_pdf_2(download_pack_files):
    with zipfile.ZipFile(resource_dir, "r") as myzip:
        with myzip.open('files/examplePdf.pdf') as myfile:
            reader = PdfReader(myfile)
            assert "рисунки" in reader.pages[1].extract_text()
            assert len(reader.pages) == 59


def test_package_excel_3(download_pack_files):
    with zipfile.ZipFile(resource_dir, "r") as myzip:
        with myzip.open('files/exampleXLSX.xlsx') as myfile:
            data = []
            workbook = load_workbook(myfile)
            sheet = workbook.active
            for row in sheet.iter_rows(values_only=True):
                data.append(row)
            assert "Оборудование для бизнеса" in data[1]
            assert len(data) == 3


def test_package_csv_4(download_pack_files):
    with zipfile.ZipFile(resource_dir, "r") as myzip:
        with myzip.open('files/exampleCSV.csv', 'r') as csv_file:
            csv_reader = list(csv.reader(TextIOWrapper(csv_file, 'utf-16')))
            assert "150000;2016-01-01;Крис Riley;trailhead9.ub20k5i9t8ou@example.com" in csv_reader[1]
            assert len(csv_reader) == 19
